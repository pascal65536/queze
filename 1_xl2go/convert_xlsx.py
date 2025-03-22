import os
import uuid
from behoof import load_json, save_json, get_md5hash, calculate_md5
import openpyxl
import datetime
import shutil
import string
import logging

logging.basicConfig(
    level=logging.INFO,
    filename="convert_xlsx.log",
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)


def read_xlsx(folder_name, file_name):
    filename = os.path.join(folder_name, file_name)
    workbook = openpyxl.load_workbook(filename)

    res_dct = {
        'title': None, 
        'rules': None,
        'date': datetime.datetime.now().isoformat(), 
        'mashup': list(), 
        'tests': list(),
    }
    for worksheet in workbook:
        if worksheet.title.strip().lower() == 'Описание'.lower():
            name = None
            for row in worksheet.iter_rows():
                res = [c.value for c in row if c.value != None]
                if not res:
                    continue
                elif res[0].strip().lower() in ["title", "rules", "date", "mashup"]:
                    name = res[0].strip().lower()
                    continue
                if not name:
                    continue
                if name == 'mashup':
                    res_dct[name].append({res[0].replace(' ', ''): res[1]})
                else:
                    res_dct[name] = res[0]
            if not res_dct['title']:
                logging.info(f'{worksheet.title=}, {filename=}, Не найдено название теста')
                exit()
            if not res_dct['rules']:
                logging.info(f'{worksheet.title=}, {filename=}, Не найдены правила теста')
                exit()
            if not res_dct['mashup']:
                logging.info(f'{worksheet.title=}, {filename=}, Не найден состав вопросов')
                exit()
                
        else:
            key = None
            this_dct = dict()
            local_dct = dict()
            worksheet_title = worksheet.title.strip().replace(' ', '')
            if worksheet_title not in [y for z in res_dct['mashup'] for y in z]:
                logging.info('Не найдено:', worksheet_title)
                exit()
            this_dct.setdefault(worksheet_title, list())
            char = 0
            for row in worksheet.iter_rows():
                res = [c.value for c in row if c.value != None]
                if not res:
                    continue

                if isinstance(res[0], str) and res[0].lower().strip() == "Question".lower():
                    if local_dct:
                        local_dct["question"] = "\n".join(local_dct["question"])
                        local_dct = dict()
                    key = "question"
                    continue
                elif isinstance(res[0], str) and res[0].lower().strip() == "Answer".lower():
                    key = "answers"
                    continue
                elif isinstance(res[0], str) and res[0].lower().strip() == "Images".lower():
                    key = "images"
                    continue
                                
                local_dct.setdefault(key, list())
                if key == "answers":
                    try:
                        local_dct[key].append({"text": res[0], "weight": res[1]})
                        local_dct['char'] = string.ascii_lowercase[char]
                    except Exception as e:
                        logging.info(f'{filename=}, {worksheet.title=}, {key=}, {res=}')
                        logging.info(e)
                        exit()
                    char += 1
                    this_dct[worksheet_title].append(local_dct)
                    local_dct = dict()
                elif key == "images":
                    images = res[0].strip()
                    inp = os.path.join(folder_name, images)
                    if not os.path.exists(inp):
                        logging.info('Не найдено:', inp)
                        exit()
                    ext = images.split(".")[-1].lower()
                    secret_filename = f"{calculate_md5(inp)}.{ext}"
                    one = get_md5hash(file_name)
                    output_path = os.path.join('output', one)
                    if not os.path.exists(output_path):
                        os.makedirs(output_path)
                    outp = os.path.join(output_path, secret_filename)
                    shutil.copy(inp, outp)
                    local_dct[key] = os.path.join(one, secret_filename)
                else:
                    local_dct[key].extend(res)
                
            res_dct['tests'].append(this_dct)
    return res_dct
           

if __name__ == "__main__":
    logging.info('Конвертация XLSX в JSON')
    input_folder = 'input'
    if not os.path.exists(input_folder):
        os.makedirs(input_folder)     

    output_folder = 'output'
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    collection = 'collection.json'
    collection_dct = load_json(output_folder, collection)

    for filename in os.listdir(input_folder):
        file_path = os.path.join(input_folder, filename)
        if not os.path.isfile(file_path):
            continue
        if '~' in filename:
            continue
        if filename.endswith(".xlsx"):
            one = get_md5hash(filename)
            quest = read_xlsx(input_folder, filename)
            collection_dct.update({one: quest})

    save_json(output_folder, collection, collection_dct)
