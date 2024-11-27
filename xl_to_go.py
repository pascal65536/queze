import os
import json
import openpyxl


def json_to_xlsx(data, filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet_num = 0

    for key, value in data.items():
        sheet_lst = list()
        for question in value:
            sheet_lst.append(["Question"])
            for q in question["question"].split("\n"):
                sheet_lst.append([q])
            sheet_lst.append(["Answer", "Value"])
            for answer in question["answers"]:
                sheet_lst.append([answer["text"], answer["weight"]])
            sheet_lst.append([""])

        if sheet_num:
            sheet = wb.create_sheet(title=key)
        sheet.title = key
        for row, sheets in enumerate(sheet_lst):
            for col, s in enumerate(sheets):
                sheet.cell(row=row + 1, column=col + 1, value=s)
        sheet_num += 1

        wb.save(filename)


def read_xlsx(folder_name, file_name):
    filename = os.path.join(folder_name, file_name)
    workbook = openpyxl.load_workbook(filename)

    res_dct = dict()
    for worksheet in workbook:
        i = 0
        key = None
        local_dct = dict()
        res_dct.setdefault(worksheet.title, list())
        for row in worksheet.iter_rows():
            res = [c.value for c in row if c.value != None]
            if not res:
                continue

            if res[0] == "Question":
                if local_dct:
                    local_dct["question"] = "\n".join(local_dct["question"])
                    res_dct[worksheet.title].append(local_dct)
                    local_dct = dict()
                i += 1
                key = "question"
                continue

            elif res[0] == "Answer":
                key = "answers"
                continue

            local_dct.setdefault(key, list())
            if key == "answers":
                local_dct[key].append({"text": res[0], "weight": res[1]})
            else:
                local_dct[key].extend(res)
    return res_dct


def read_data(folder_name, file_name):
    if not os.path.exists(folder_name):
        os.mkdir(folder_name)
    filename = os.path.join(folder_name, file_name)
    if not os.path.exists(filename):
        with open(filename, "w") as f:
            json.dump(dict(), f, ensure_ascii=True)
    with open(filename, encoding="utf-8") as f:
        load_dct = json.load(f)
    return load_dct


def write_data(folder_name, file_name, save_dct):
    if not os.path.exists(folder_name):
        os.mkdir(folder_name)
    filename = os.path.join(folder_name, file_name)
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(save_dct, f, ensure_ascii=False, indent=4)


def convert_dir(folder_name, json_filename):
    """
    Обновление библиотеки тестов
    """
    collect = read_data(folder_name, json_filename)
    
    output_folder_name = os.path.join(folder_name, "output")
    if not os.path.exists(output_folder_name):
        os.mkdir(output_folder_name)
    for quest in collect:
        quest_name = os.path.join(output_folder_name, f"{quest}.xlsx")
        if os.path.exists(quest_name):
            os.remove(quest_name)
        json_to_xlsx(collect[quest], quest_name)
   
    input_folder_name = os.path.join(folder_name, "input")  
    if not os.path.exists(input_folder_name):
        os.mkdir(input_folder_name)      
    for root, _, files in os.walk(input_folder_name):
        for file_name in files:
            *fn, fe = file_name.split(".")
            if fe != "xlsx":
                continue
            key = ".".join(fn)
            quest = read_xlsx(input_folder_name, file_name)
            if sum([len(quest[q]) for q in quest]):
                collect.update({key: quest})
            else:
                if key in collect:
                    del collect[key]

    write_data(folder_name, collection, collect)

def get_xlsx(filename):
    workbook = openpyxl.load_workbook(filename)
    for worksheet in workbook:
        # for row in worksheet.iter_rows():
        #     print(row)
        print(worksheet)
            


if __name__ == "__main__":
    folder_name = "queze/tests"
    collection = "collection.json"
    xl = "tests/input/Проверочная работа по Информатике.xlsx"
    get_xlsx(xl)
    # convert_dir(folder_name, collection)
