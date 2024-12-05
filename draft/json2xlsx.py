import behoof as utils
import pandas as pd
import os
import openpyxl


def json_to_excel(json_file, excel_file, folder_name = "data"):
    """
    Converts a JSON file to an Excel file.
    """
    if not os.path.exists(folder_name):
        os.mkdir(folder_name)
    excel_filename = os.path.join(folder_name, excel_file)
    data = utils.load_json(folder_name, json_file)
    with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
        root = list()
        root.append("Название")
        root.append(data.get("title", ""))
        root.append("Правила")
        root.append(data.get("rules", ""))
        root.append("Состав вопросов")
        for mashup in data.get("mashup", list()):
            root.append(mashup)
        df = pd.DataFrame(root, columns=[None])
        df.to_excel(writer, sheet_name="Описание", index=False)
        for topic in data["tests"]:
            title = topic["title"]
            rows = []
            rows = list()
            for question in topic.get("questions", list()):
                rows.append(["Question", None])
                for q in question.get("question", list()):
                    rows.append([q.strip(), None])
                images = question.get("images", list())
                if images:
                    rows.append(["Image", None])
                    for i in images:
                        rows.append([i.strip(), None])
                rows.append(["Answer", "Weight"])
                for a in question.get("answer", list()):
                    rows.append([a["text"], a["weight"]])
                rows.append("")
            df = pd.DataFrame(rows, columns=[None, None])
            df.to_excel(writer, sheet_name=title[:31], index=False)


def json_to_json(json_file, out_json):
    """
    Функция для преобразования JSON
    """
    data = utils.load_json("queze", json_file)
    topic_title = data["title"]
    topic_rules = data["rules"]
    topic_mashup = data["mashup"]

    topic_dct = dict()
    topic_dct.setdefault("title", topic_title)
    topic_dct.setdefault("rules", topic_rules)
    topic_dct.setdefault("mashup", topic_mashup)
    topic_dct.setdefault("tests", list())
    topic = topic_dct["tests"]
    for data in data["tests"]:
        title = data["title"]
        questions = data["questions"]
        rows = list()
        for question in questions:
            row_dct = dict()
            row_dct.setdefault("image", None)
            row_dct.setdefault("question", list())
            for q in question["question"].split("\n"):
                row_dct["question"].append(q)

            row_dct.setdefault("answer", list())
            if "answer" in question and "points" in question:
                row_dct["answer"].append(
                    {
                        "text": question.get("answer"),
                        "weight": question.get("points"),
                    }
                )

            for answer in question.get("answers", list()):
                if "text" in answer and "weight" in answer:
                    row_dct["answer"].append(
                        {
                            "text": question.get("answer"),
                            "weight": question.get("points"),
                        }
                    )

            rows.append(row_dct)
        topic.append({"title": title, "questions": rows})
    utils.save_json("queze", out_json, topic_dct)


def excel_to_json(excel_file, json_file, folder_name = "data"):
    """
    Функция для преобразования Excel в JSON
    """
    if not os.path.exists(folder_name):
        os.mkdir(folder_name)
    excel_filename = os.path.join(folder_name, excel_file)
    workbook = openpyxl.load_workbook(excel_filename)
    res_dct = {"title": list(), "rules": list(), "mashup": list(), "tests": list()}

    for worksheet in workbook:
        if worksheet.title.lower() == "описание":
            key = None
            for row in worksheet.iter_rows():
                res = [c.value for c in row if c.value != None]
                if not res:
                    continue
                elif "Название" in res:
                    key = "title"
                elif "Правила" in res:
                    key = "rules"
                elif "Состав вопросов" in res:
                    key = "mashup"
                else:
                    res_dct[key].append(res[0])
        else:
            local_lst = list()
            key = None
            this_dct = {"question": list(), "answers": list()}
            for row in worksheet.iter_rows():
                res = [c.value for c in row if c.value != None]
                if not res:
                    continue
                if res[0] == "Question":
                    key = "question"
                    if this_dct["question"] and this_dct["answers"]:
                        local_lst.append(this_dct)
                    this_dct = {"question": list(), "answers": list()}
                elif res[0] == "Answer":
                    key = "answers"
                else:
                    if key == "answers":
                        value = {"text": res[0], "weight": res[1]}
                    else:
                        value = res[0]
                    this_dct[key].append(value)
            if this_dct["question"] and this_dct["answers"]:
                local_lst.append(this_dct)
            res_dct["tests"].append({"title": worksheet.title, "questions": local_lst})
    data = utils.load_json(folder_name, json_file)
    data[excel_file] = res_dct
    utils.save_json(folder_name, json_file, data)


if __name__ == "__main__":
    out_json = "collection.json"
    excel_file = "collection.xlsx"
    # json_to_json(json_file, out_json)
    # json_to_excel(out_json, excel_file)
    excel_to_json(excel_file, out_json)
