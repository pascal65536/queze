import os
import uuid
from behoof import load_json, save_json, get_md5hash, calculate_md5
import openpyxl
import datetime
import shutil


def json_to_xlsx(data, filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet_num = 0

    for key, value in data.items():
        sheet_lst = list()
        for question in value:
            sheet_lst.append(["Question"])
            for q in question["question"].split("\n"):
                sheet_lst.append([q])
            sheet_lst.append(["Answer", "Value"])
            for answer in question["answers"]:
                sheet_lst.append([answer["text"], answer["weight"]])
            sheet_lst.append([""])

        if sheet_num:
            sheet = wb.create_sheet(title=key)
        sheet.title = key
        for row, sheets in enumerate(sheet_lst):
            for col, s in enumerate(sheets):
                sheet.cell(row=row + 1, column=col + 1, value=s)
        sheet_num += 1

        wb.save(filename)
           

if __name__ == "__main__":
    print('Конвертация JSON в XLSX')
    output_folder = 'output'
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    input_folder = 'data'
    if not os.path.exists(input_folder):
        os.makedirs(input_folder)     

    collection = 'collection.json'
    collection_dct = load_json(output_folder, collection)

    for k, collection in collection_dct.items():
        print(k)
        title = collection.get('title')
        rules = collection.get('rules')
        date = collection.get('date', datetime.datetime.now().isoformat())
        mashup = collection.get('mashup', list())
        tests = collection.get('tests', list())
        print(title)
        print(rules)
        print(date)
        print(mashup)
        # print(tests)

    # for filename in os.listdir(input_folder):
    #     file_path = os.path.join(input_folder, filename)
    #     if not os.path.isfile(file_path):
    #         continue
    #     if '~' in filename:
    #         continue
    #     if filename.endswith(".xlsx"):
    #         one = get_md5hash(filename)
    #         quest = read_xlsx(input_folder, filename)
    #         collection_dct.update({one: quest})

    # save_json(output_folder, collection, collection_dct)
